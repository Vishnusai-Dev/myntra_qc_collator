# app.py
"""
Streamlit app implementing Option B:
- Inserts one combined core-bucket column (Compliance, Content, ...) BEFORE original check columns.
- Keeps all original non-Reason columns in exactly their original order.
- Puts reason messages both in the specific check column (if matched) AND concatenated in the combined bucket column.
- Exports a 2-row header: Row1 = bucket names, Row2 = check column names.
- Groups and collapses Reason columns at the end of the exported workbook.
"""
import streamlit as st
import pandas as pd
import re
from collections import defaultdict, OrderedDict
from io import BytesIO
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# optional fuzzy matching
try:
    from rapidfuzz import process, fuzz
    RAPIDFUZZ = True
except Exception:
    RAPIDFUZZ = False

st.set_page_config(page_title="Myntra QC Collator (Option B)", layout="wide")

CORE_ORDER = ["Compliance", "Content", "Formatting", "Image Formatting", "Image", "Size", "Image Sequence"]
DEFAULT_FUZZ = 82

def normalize(s):
    if s is None:
        return ""
    s = str(s).strip()
    s = re.sub(r'[\t\r\n]+', ' ', s)
    s = re.sub(r'\s+', ' ', s)
    s = s.strip(" .:-–—")
    return s.lower()

def split_bullets(text):
    if pd.isna(text) or text is None:
        return []
    s = str(text)
    parts = re.split(r'\n|•|◦|·|-{2,}|—|–', s)
    return [p.strip() for p in parts if p.strip()]

def parse_title_msg(bullet):
    if not bullet:
        return ("","")
    parts = re.split(r'\s*[:\-—–]\s*', bullet, maxsplit=1)
    if len(parts)==2:
        return parts[0].strip(), parts[1].strip()
    return parts[0].strip(), ""

def build_grouped_excel_2rowheader(style_col, combined_cols, original_cols, original_row_values, reason_columns, mapping_audit_df):
    """
    Create workbook with:
    - Row1: bucket names (for combined cols then for each original col its bucket)
    - Row2: column names (combined col names and original col names)
    - Then data rows from original_row_values (list of dicts)
    - At the end append reason columns (grouped & hidden)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "cleaned_output"

    # Build Row1 and Row2 headers
    row1 = []
    row2 = []
    # style id column(s)
    row1.append("")  # no bucket above styleid
    row2.append(style_col)

    # combined bucket columns (one per core in CORE_ORDER) - bucket name on row1, label on row2 (we use same name as row1)
    for core in combined_cols:
        row1.append(core)
        row2.append(core)  # column name same as bucket (optionally "Compliance (Combined)" if you want)

    # For original columns, look up bucket label from mapping_audit_df if exists, else blank
    col_to_bucket = {}
    if mapping_audit_df is not None and not mapping_audit_df.empty:
        # mapping_audit_df expected columns: Header, CoreBucket
        for _, r in mapping_audit_df.iterrows():
            hdr = str(r.get("Header","")).strip()
            bucket = str(r.get("CoreBucket","")).strip()
            col_to_bucket[hdr] = bucket

    for col in original_cols:
        bucket = col_to_bucket.get(col, "")
        row1.append(bucket)
        row2.append(col)

    # write header rows
    ws.append(row1)
    ws.append(row2)

    # write data rows (in same order)
    for rd in original_row_values:
        row = []
        # styleid
        row.append(rd.get(style_col,""))
        # combined columns first
        for core in combined_cols:
            row.append(rd.get(f"__COMBINED__{core}", ""))
        # then original columns
        for col in original_cols:
            row.append(rd.get(col, ""))
        ws.append(row)

    # Now append reason columns as separate sheet area to the right and then set them hidden (grouped)
    # We'll add them as additional columns in the same sheet after current columns, to allow Excel grouping
    current_col_count = len(row1)
    # append reason headers to header rows
    for rc in reason_columns:
        ws.cell(row=1, column=current_col_count+1).value = ""  # top row: we leave blank or could put "Reason"
        ws.cell(row=2, column=current_col_count+1).value = rc
        current_col_count += 1

    # fill reason values per row
    # original_row_values are in the same order, so for each row index i (0-based) write reason values
    for i, rd in enumerate(original_row_values):
        excel_row_idx = 3 + i  # data starts at row 3
        col_idx = len(row1) + 1
        for rc in reason_columns:
            ws.cell(row=excel_row_idx, column=col_idx).value = rd.get(rc, "")
            col_idx += 1

    # set outline level and hide reason columns so they appear collapsed (+) in Excel
    start_reason_col = len(row1) + 1
    end_reason_col = current_col_count
    for cidx in range(start_reason_col, end_reason_col+1):
        letter = get_column_letter(cidx)
        try:
            ws.column_dimensions[letter].outlineLevel = 1
            ws.column_dimensions[letter].hidden = True
        except Exception:
            pass

    return wb

# UI
st.title("Myntra QC Collator — Option B (Combined bucket column + per-check messages)")

uploaded = st.file_uploader("Upload Output Excel (.xlsx)", type=["xlsx"])
col1, col2, col3 = st.columns(3)
with col1:
    style_col = st.text_input("styleId column name", value="styleid")
with col2:
    summary_col = st.text_input("Failure Summary column name", value="Failure Summary")
with col3:
    report_col = st.text_input("Failure Report column name", value="Failure Report")

# load bundled mapping.xlsx
mapping_path = os.path.join(os.path.dirname(__file__), "mapping.xlsx")
if not os.path.exists(mapping_path):
    st.error("mapping.xlsx missing in repo. Add mapping.xlsx and redeploy.")
    st.stop()

map_df = pd.read_excel(mapping_path).fillna("")
# normalize expected columns
if len(map_df.columns) >= 2:
    map_df = map_df.iloc[:, :2]
    map_df.columns = ["Header", "CoreBucket"]
else:
    map_df = pd.DataFrame(columns=["Header","CoreBucket"])

# show editable mapping (best-effort)
editor_df = None
try:
    editor_df = st.data_editor(map_df, num_rows="dynamic")
except Exception:
    try:
        editor_df = st.experimental_data_editor(map_df, num_rows="dynamic")
    except Exception:
        st.write("Mapping editor not available; using bundled mapping.")
        editor_df = map_df.copy()

# Build mapping dict normalized -> bucket and header name set
mapping_norm_to_bucket = {}
all_mapping_headers = []
for _, r in editor_df.iterrows():
    hdr = str(r.get("Header","")).strip()
    bucket = str(r.get("CoreBucket","")).strip() or ""
    if hdr:
        mapping_norm_to_bucket[normalize(hdr)] = bucket
        all_mapping_headers.append(hdr)

# Precompute normalized mapping keys list (for fuzzy)
norm_keys = list(mapping_norm_to_bucket.keys())

if uploaded is None:
    st.info("Upload your Output Excel to process.")
    st.stop()

# Read uploaded workbook (first sheet)
try:
    xls = pd.read_excel(uploaded, sheet_name=None, dtype=str)
    first_sheet_name = list(xls.keys())[0]
    df = xls[first_sheet_name].fillna("")
except Exception as e:
    st.error("Failed to read uploaded Excel: " + str(e))
    st.stop()

# Identify reason columns (headers containing 'Reason')
all_cols = list(df.columns)
reason_cols = [c for c in all_cols if re.search(r'\breason\b', str(c), flags=re.IGNORECASE)]

# Build original non-reason columns in original order
original_non_reason_cols = [c for c in all_cols if c not in reason_cols]

# We will keep all original_non_reason_cols exactly in the same order
# We'll create combined bucket columns (one per CORE_ORDER) and place them before original columns

# Build a fast lookup to match a reason title to a specific original check column name if present
# We'll use normalized matching against original_non_reason_cols too
orig_norm_to_col = {normalize(c): c for c in original_non_reason_cols}

# function to map a reason title to a bucket:
def map_title_to_bucket(title, fuzz_threshold=DEFAULT_FUZZ):
    n = normalize(title)
    # exact normalized
    if n in mapping_norm_to_bucket and mapping_norm_to_bucket[n]:
        return mapping_norm_to_bucket[n], "exact"
    # substring match against mapping headers (case-insensitive)
    for hdr in all_mapping_headers:
        if hdr and hdr.lower() in title.lower():
            return mapping_norm_to_bucket.get(normalize(hdr), ""), "substr_maphdr"
    # fuzzy match if available
    if RAPIDFUZZ and norm_keys:
        best = process.extractOne(n, norm_keys, scorer=fuzz.token_sort_ratio)
        if best and best[1] >= fuzz_threshold:
            return mapping_norm_to_bucket[best[0]], f"fuzzy:{best[1]}"
    return "", "unmapped"

# We'll also try to find the original column that corresponds to a particular reason title
def find_matching_orig_col(title, fuzz_threshold=DEFAULT_FUZZ):
    n = normalize(title)
    if n in orig_norm_to_col:
        return orig_norm_to_col[n], "exact_col"
    # try substring in original column names
    for col in original_non_reason_cols:
        if col and col.lower() in title.lower():
            return col, "substr_col"
        if title.lower() in col.lower():
            return col, "substr_col2"
    if RAPIDFUZZ and orig_norm_to_col:
        best = process.extractOne(n, list(orig_norm_to_col.keys()), scorer=fuzz.token_sort_ratio)
        if best and best[1] >= DEFAULT_FUZZ:
            return orig_norm_to_col[best[0]], f"fuzzy_col:{best[1]}"
    return None, "no_col"

# Prepare output row container list of dicts
output_rows = []
mapping_audit = []  # for mapping_used sheet

# For each row, parse bullets and populate combined buckets & column-specific cells
for _, row in df.iterrows():
    # initialize output row with original non-reason column values
    out_row = {c: row.get(c, "") for c in original_non_reason_cols}
    # initialize combined fields
    for core in CORE_ORDER:
        out_row[f"__COMBINED__{core}"] = ""
    # parse summary and report bullets
    summary_bullets = split_bullets(row.get(summary_col, ""))
    report_bullets = split_bullets(row.get(report_col, ""))
    # parse into (title,msg) lists
    summary_parsed = [parse_title_msg(b) for b in summary_bullets]
    report_parsed = [parse_title_msg(b) for b in report_bullets]
    # build a report_map: normalized title -> list(messages)
    report_map = defaultdict(list)
    for rt, rm in report_parsed:
        key = normalize(rt) if rt else normalize(rm[:40])
        report_map[key].append(rm or "")
    # process summary parsed bullets primarily
    bucket_accumulator = defaultdict(list)  # bucket -> list of messages
    for title, s_msg in summary_parsed:
        if not title:
            continue
        # prefer a detailed message from report if possible
        repmsg = ""
        key = normalize(title)
        if key in report_map and report_map[key]:
            repmsg = report_map[key][0]
        else:
            # try substring match in report_parsed
            for rt, rm in report_parsed:
                hay = (rt or "") + " " + (rm or "")
                if title and title.lower() in hay.lower():
                    repmsg = rm or ""
                    break
        final_msg = repmsg if repmsg else s_msg

        # map title to bucket
        bucket, method = map_title_to_bucket(title)
        mapping_audit.append({"Title": title, "Bucket": bucket or "Unmapped", "Method": method})
        if not bucket:
            bucket = "Unmapped"
        # append to combined bucket accumulator
        bucket_accumulator[bucket].append(f"{title}: {final_msg}".strip())

        # attempt to find matching original column and write message there (overwriting previous value)
        matched_col, col_method = find_matching_orig_col(title)
        if matched_col:
            # put final_msg into that column's cell
            out_row[matched_col] = f"{title}: {final_msg}"
            mapping_audit[-1]["MatchedColumn"] = matched_col
            mapping_audit[-1]["ColMatchMethod"] = col_method
        else:
            mapping_audit[-1]["MatchedColumn"] = ""
            mapping_audit[-1]["ColMatchMethod"] = col_method

    # process any report bullets not covered above
    # (if report bullet not mapped to summary, we still map it)
    for rt, rm in report_parsed:
        key = normalize(rt)
        # skip if this report message already used in above mapping (we used report_map entries above)
        # simple heuristic: if rm is empty and rt exists in summary_parsed titles, skip
        already_handled = False
        for t, _ in summary_parsed:
            if normalize(t) == key:
                already_handled = True
                break
        if already_handled:
            continue
        title = rt or (rm[:40] if rm else "Unknown")
        final_msg = rm or ""
        bucket, method = map_title_to_bucket(title)
        mapping_audit.append({"Title": title, "Bucket": bucket or "Unmapped", "Method": f"report_{method}"})
        if not bucket:
            bucket = "Unmapped"
        bucket_accumulator[bucket].append(f"{title}: {final_msg}")
        matched_col, col_method = find_matching_orig_col(title)
        if matched_col:
            out_row[matched_col] = f"{title}: {final_msg}"
            mapping_audit[-1]["MatchedColumn"] = matched_col
            mapping_audit[-1]["ColMatchMethod"] = col_method
        else:
            mapping_audit[-1]["MatchedColumn"] = ""
            mapping_audit[-1]["ColMatchMethod"] = col_method

    # fill combined columns in out_row
    for core in CORE_ORDER:
        msgs = bucket_accumulator.get(core, [])
        out_row[f"__COMBINED__{core}"] = "\n".join(msgs) if msgs else ""

    # also store original reason columns' values (we'll keep and export them at the end)
    for rc in reason_cols:
        out_row[rc] = row.get(rc,"")

    # store style col
    out_row[style_col] = row.get(style_col, "")

    # append to output rows
    output_rows.append(out_row)

# Build final ordered lists for export
# Combined columns first (in CORE_ORDER)
combined_cols = CORE_ORDER.copy()

# Original columns: keep original_non_reason_cols order
original_cols = original_non_reason_cols.copy()

# We'll build a mapping audit dataframe to put in the workbook for review
mapping_audit_df = pd.DataFrame(mapping_audit)

# Prepare mapping header info for row1 above original columns:
# For each original column, find its bucket using mapping_norm_to_bucket or heuristics
orig_col_buckets = []
for col in original_cols:
    n = normalize(col)
    bucket = mapping_norm_to_bucket.get(n, "")
    if not bucket:
        # try substring of mapping headers
        found = ""
        for hdr in all_mapping_headers:
            if hdr and hdr.lower() in col.lower():
                found = mapping_norm_to_bucket.get(normalize(hdr), "")
                break
        bucket = found or ""
    orig_col_buckets.append(bucket)

# Prepare original_row_values as list of dicts in same order as output_rows
# (already done: output_rows list)
# Now create workbook with 2-row header and grouped reason columns

wb = build_grouped_excel_2rowheader(style_col, combined_cols, original_cols, output_rows, reason_cols, editor_df)

# convert to bytes
buf = BytesIO()
wb.save(buf)
buf.seek(0)

st.success("Prepared output — download the Excel below.")
st.download_button(
    "Download 2-row-header grouped workbook (Option B)",
    data=buf.getvalue(),
    file_name=f"myntra_qc_optionB_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

st.info("Notes: Combined bucket columns appear after styleId (one column per core). Each original column retains its original position and name. Reason columns are appended and grouped (collapsed).")
